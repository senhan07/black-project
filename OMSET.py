
# import module
from lib2to3.pgen2.pgen import DFAState
from os import system, name
from traceback import print_stack
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import os

pwd = os.getcwd()

def clear():
    if name == 'nt':
        _ = system('cls')
    else:
        _ = system('clear')
clear()



print("OMSET SALES EXTRACT TOOLS")
filename = input("File Excel:")
#filename = "C:/Users/losen/Desktop/RAW/JAN-22.xls"

sales_array = [
    'AWI SUHARTO',
    'SINATRA',
    'ANDRY',
    'ANDRE KS',
    'P.MAR',
    'YOYOK',
    'WENDY',
    'NURSAMSI',
    'RULI',
    'MAT',
    'ROBY',
    'CANDRA',
    'JULIAN',
    'BAMBANG',
    'STEVEN',
    'GEOVANO',
    'HENDRA'
]

print("")
print("NAMA SALES, PUSHON, CAHAYA, DEXICON, CAMPUR, TOTAL")

#while(True):
for sales in sales_array:
    total = pd.read_excel(filename, skiprows=[0], usecols = "B,H,I,L,Y")
    pushon = pd.read_excel(filename, skiprows=[0], usecols = "B,H,I,L,Y")
    cahaya = pd.read_excel(filename, skiprows=[0], usecols = "B,H,I,L,Y")
    dexicon = pd.read_excel(filename, skiprows=[0], usecols = "B,H,I,L,Y")
    #clear()
    #print("OMSET SALES EXTRACT TOOLS")
    #sales = input("Nama Sales:")

    total = total[total['Salesman'].str.contains(sales, na=False)]
    total_total = total['Total'].sum()

    pushon = pushon[pushon['Salesman'].str.contains(sales, na=False)]
    pushon = pushon[pushon['Brand'].str.contains('PUSH O|LUMMEN', na=False)]
    total_pushon = pushon['Total'].sum()

    cahaya = cahaya[cahaya['Salesman'].str.contains(sales, na=False)]
    cahaya = cahaya[cahaya['Brand'].str.contains('CAHAYA|SZMR|PANCAR|HAKAMI|FROZEN', na=False)]
    total_cahaya = cahaya['Total'].sum()

    dexicon = dexicon[dexicon['Salesman'].str.contains(sales, na=False)]
    dexicon = dexicon[dexicon['Brand'].str.contains('CROWN|DEXICO|ENJE|ENJE O|HERCUL|KINGWO|KSRAN|LEXICO|NJ|UNITED|YUKI|VENUS|VENUSI|VENUSIA|DEXICON|GRONCO|KINGWON|LEXICON|LUXURI', na=False)]
    total_dexicon = dexicon['Total'].sum()

    #print(total)
    #print("")
    #print(pushon)
    #print("")
    #print(cahaya)
    #print("")
    #print(dexicon)
    #print("")
    #print("")
    #print(sales)
    #print("-------------------------")
    #print(f"TOTAL: {total_total:,}")
    #print("-------------------------")
    #print(f"PUSHON: {total_pushon:,}")
    #print(f"CAHAYA: {total_cahaya:,}")
    #print(f"DEXICON: {total_dexicon:,}")
    total_campur = total_total-(total_pushon+total_cahaya+total_dexicon)
    #print("CAMPUR: " + f"{total_campur:,}")
    #print("")

    #SET DATAFRAME
    df_marks = pd.DataFrame({'PUSHON': [total_pushon],
     'CAHAYA': [total_cahaya],
     'DEXICON': [total_dexicon],
     'CAMPUR': [total_campur]},
     index = [sales])
	
    print(sales,',',total_pushon,',',total_cahaya,',',total_dexicon,',',total_campur,',',total_total)
    #wb = load_workbook(r'output.xlsx')
    #writer = pd.ExcelWriter(r'output.xlsx', engine='openpyxl')
    #df_marks.to_excel(writer, startrow = 2)
    #wb.save(r'output.xlsx')
    #print("done")



    # SAVE TO EXCEL
    #book = load_workbook(pwd + '\output.xlsx')
    #writer = pd.ExcelWriter(pwd + '\output.xlsx', engine='openpyxl')
    ##writer.book = book
    #i = 0
    #while i < 5:
    # print(i)
    # i += 3
    #df_marks.to_excel(writer, startrow=i)
    #writer.save()
    #print('DataFrame is written successfully to Excel File.')
    #print(df_marks)